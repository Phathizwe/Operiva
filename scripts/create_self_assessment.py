#!/usr/bin/env python3.11
"""
Create Owner-Manager Self-Assessment Worksheet (Capability Compass)
Based on the research blueprint - scenario-based assessment across 5 capability areas
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import RadarChart, Reference
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# Remove default sheet
wb.remove(wb.active)

# Define Operiva colors
NAVY = "1C2A4A"
GREEN = "27AE60"
LIGHT_BLUE = "E8F4F8"
LIGHT_GREEN = "E8F5E9"
YELLOW = "FFF9C4"
RED = "FFEBEE"

# Define styles
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
subheader_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
subheader_font = Font(color="FFFFFF", bold=True, size=11)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# SHEET 1: INSTRUCTIONS
# ============================================================================
ws_instructions = wb.create_sheet("📖 Instructions")

ws_instructions['A1'] = "Owner-Manager Self-Assessment: The Capability Compass"
ws_instructions['A1'].font = Font(size=18, bold=True, color=NAVY)
ws_instructions.merge_cells('A1:F1')

ws_instructions['A3'] = "Welcome to Your Business Capability Assessment"
ws_instructions['A3'].font = Font(size=14, bold=True)
ws_instructions.merge_cells('A3:F3')

instructions_text = """
This self-assessment tool helps you identify your strengths and gaps across 5 critical business capabilities:

1️⃣ Financial Management - Cash flow, pricing, bookkeeping
2️⃣ Strategic Planning - Vision, goal-setting, market analysis
3️⃣ Leadership & People - Team management, delegation, culture
4️⃣ Marketing & Sales - Customer acquisition, branding, digital presence
5️⃣ Operations & Technology - Processes, efficiency, digital tools

HOW IT WORKS:
• Each capability area has 10 scenario-based questions
• Read each scenario carefully - they're based on real SA SMME situations
• Choose the response that best matches what YOU would do
• Be honest! This is for YOUR benefit, not a test
• Your scores will automatically calculate in the Dashboard tab

SCORING:
• 0-3 points = Needs Urgent Attention (Red)
• 4-6 points = Developing (Yellow)
• 7-8 points = Competent (Light Green)
• 9-10 points = Strong (Dark Green)

TIME: 20-30 minutes to complete all 50 questions

START: Go to the "1️⃣ Financial" tab and begin!
"""

ws_instructions['A5'] = instructions_text
ws_instructions['A5'].alignment = left_align
ws_instructions.merge_cells('A5:F25')
ws_instructions.row_dimensions[5].height = 400

# Set column widths
for col in range(1, 7):
    ws_instructions.column_dimensions[get_column_letter(col)].width = 15

# ============================================================================
# SHEET 2-6: ASSESSMENT QUESTIONS (5 Capability Areas)
# ============================================================================

capabilities = [
    {
        "name": "1️⃣ Financial",
        "full_name": "Financial Management",
        "scenarios": [
            ("Cash Flow Crisis", "It's month-end and you don't have enough cash to pay suppliers. What do you do?",
             [("Ask family for a loan", 0), ("Review which invoices are overdue and chase customers", 2), ("Delay paying suppliers without telling them", 0), ("Check your cash flow forecast to see when money is coming in", 3)]),
            
            ("Pricing Decision", "A competitor drops their prices by 20%. Your customers are asking for discounts. What's your response?",
             [("Match their prices immediately", 0), ("Calculate your costs first, then decide", 3), ("Offer a discount to keep customers happy", 1), ("Ignore it and hope customers stay loyal", 0)]),
            
            ("Bookkeeping Backlog", "Your bookkeeping is 3 months behind. Tax season is approaching. What do you prioritize?",
             [("Hire a bookkeeper immediately", 2), ("Do it yourself over the weekend", 1), ("Set up a simple system and update weekly going forward", 3), ("Wait until SARS sends a letter", 0)]),
            
            ("Unexpected Expense", "Your delivery vehicle breaks down. Repair costs R15,000. How do you handle it?",
             [("Put it on credit card", 1), ("Check if you have emergency funds set aside", 3), ("Borrow from business cash", 1), ("Use money meant for supplier payments", 0)]),
            
            ("Profit vs Revenue", "Your revenue increased 30% this year, but you're not seeing more money in the bank. Why?",
             [("Customers must be stealing", 0), ("Check if costs increased or if money is tied up in stock/debtors", 3), ("Revenue doesn't matter, only cash matters", 2), ("It's normal for growing businesses", 0)]),
            
            ("VAT Registration", "Your turnover just crossed R1 million. A friend says you must register for VAT. What do you do?",
             [("Ignore it until SARS contacts you", 0), ("Research the VAT threshold and registration process", 3), ("Register immediately without understanding implications", 1), ("Ask your friend to help you", 1)]),
            
            ("Credit Terms", "A big customer wants 60-day payment terms. You usually get paid in 30 days. What do you consider?",
             [("Accept - big customers are always right", 0), ("Calculate if your cash flow can handle the delay", 3), ("Reject - you need cash now", 1), ("Accept but charge them more", 2)]),
            
            ("Financial Reports", "How often do you review your Profit & Loss statement?",
             [("What's a P&L?", 0), ("Once a year for tax", 1), ("Monthly, to track performance", 3), ("Never, I just check my bank balance", 0)]),
            
            ("Pricing Formula", "How do you set prices for your products/services?",
             [("See what competitors charge", 1), ("Calculate costs + desired profit margin", 3), ("Charge what customers will pay", 1), ("Use a price list from 3 years ago", 0)]),
            
            ("Separating Finances", "Do you have separate bank accounts for business and personal use?",
             [("No, it's all my money anyway", 0), ("Yes, and I pay myself a salary", 3), ("Sometimes I use business money for personal needs", 1), ("I have separate accounts but mix them often", 1)])
        ]
    },
    {
        "name": "2️⃣ Strategic",
        "full_name": "Strategic Planning & Vision",
        "scenarios": [
            ("Business Plan", "When did you last update your business plan?",
             [("I don't have one", 0), ("I have one in my head", 1), ("I wrote one 3 years ago", 1), ("I review and update it every 6-12 months", 3)]),
            
            ("Market Changes", "Load shedding is affecting your business. How do you respond?",
             [("Complain and hope it ends soon", 0), ("Develop a continuity plan (generator, solar, adjusted hours)", 3), ("Close during power cuts", 1), ("Just deal with it day by day", 0)]),
            
            ("Goal Setting", "What are your business goals for the next 12 months?",
             [("Survive", 0), ("Increase sales", 1), ("Specific targets: e.g., R500k revenue, 2 new products, hire 1 staff", 3), ("I don't set goals", 0)]),
            
            ("Competitor Analysis", "A new competitor opens nearby. What do you do?",
             [("Panic and drop prices", 0), ("Analyze what they offer vs what makes you different", 3), ("Ignore them", 0), ("Copy what they're doing", 1)]),
            
            ("Growth Opportunity", "A customer asks if you can supply a new product line. How do you decide?",
             [("Say yes immediately - more sales!", 1), ("Assess: Can we deliver? Does it fit our strategy? Is it profitable?", 3), ("Say no - too risky", 0), ("Ask other customers if they'd want it too", 2)]),
            
            ("SWOT Analysis", "Have you done a SWOT analysis (Strengths, Weaknesses, Opportunities, Threats) for your business?",
             [("Never heard of it", 0), ("Yes, and I use it to make decisions", 3), ("I did one once in a workshop", 1), ("I know my strengths, that's enough", 1)]),
            
            ("Pivot Decision", "Your main product isn't selling well anymore. What do you do?",
             [("Keep pushing it - customers will come back", 0), ("Research why sales dropped and test new offerings", 3), ("Shut down the business", 0), ("Copy what's trending", 1)]),
            
            ("Vision Communication", "Do your employees (if any) know your business vision and goals?",
             [("I don't have employees", 0), ("No, they just need to do their jobs", 0), ("Yes, I share and discuss goals regularly", 3), ("I mentioned it once when I hired them", 1)]),
            
            ("Risk Management", "What's your biggest business risk right now, and how are you managing it?",
             [("I don't think about risks", 0), ("I know my risks and have plans to reduce them", 3), ("I worry about everything but don't have plans", 1), ("I'll deal with problems when they happen", 0)]),
            
            ("Strategic Partnerships", "How do you approach partnerships with other businesses?",
             [("I don't partner with anyone", 0), ("I evaluate if it's mutually beneficial and aligns with my strategy", 3), ("I partner with anyone who asks", 1), ("Partnerships are for big companies", 0)])
        ]
    },
    {
        "name": "3️⃣ Leadership",
        "full_name": "Leadership & People Management",
        "scenarios": [
            ("Delegation", "You're working 70 hours a week. What do you do?",
             [("Keep working - no one can do it like me", 0), ("Identify tasks others can do and train them", 3), ("Hire someone and hope they figure it out", 1), ("Accept it - that's entrepreneurship", 0)]),
            
            ("Employee Conflict", "Two employees are fighting. How do you handle it?",
             [("Ignore it - they'll sort it out", 0), ("Meet with each privately, then together to resolve", 3), ("Fire both of them", 0), ("Tell them to stop being childish", 1)]),
            
            ("Performance Issues", "An employee is underperforming. What's your approach?",
             [("Fire them immediately", 0), ("Have a conversation to understand why and set clear expectations", 3), ("Reduce their hours/pay", 1), ("Ignore it and do their work yourself", 0)]),
            
            ("Team Motivation", "How do you keep your team motivated?",
             [("Pay them - that's enough", 1), ("Recognize good work, provide growth opportunities, clear communication", 3), ("Threaten to fire them if they slack", 0), ("I don't have time to motivate people", 0)]),
            
            ("Hiring Decision", "You need to hire someone. What's your process?",
             [("Hire the first person who applies", 0), ("Define the role, interview multiple candidates, check references", 3), ("Hire a family member", 1), ("Post on Facebook and see who responds", 1)]),
            
            ("Feedback Culture", "How often do you give feedback to your team?",
             [("Only when they mess up", 0), ("Regularly - both positive and constructive", 3), ("Never, they should know if they're doing well", 0), ("Once a year", 1)]),
            
            ("Leadership Style", "How would you describe your leadership approach?",
             [("I'm the boss, they must listen", 0), ("I lead by example and empower my team", 3), ("I'm too nice, people take advantage", 1), ("I don't think of myself as a leader", 0)]),
            
            ("Training Investment", "Do you invest in training for yourself and your team?",
             [("No, it's too expensive", 0), ("Yes, regularly - it's an investment", 3), ("Only if it's free", 1), ("I learn from YouTube", 2)]),
            
            ("Succession Planning", "If you couldn't work for 3 months, could your business survive?",
             [("No, everything depends on me", 0), ("Yes, I've documented processes and trained others", 3), ("Maybe, but it would struggle", 1), ("I've never thought about it", 0)]),
            
            ("Culture Building", "What's your business culture like?",
             [("Culture? We just work", 0), ("I actively build a positive, values-driven culture", 3), ("It's whatever happens naturally", 1), ("I don't have enough staff to worry about culture", 0)])
        ]
    },
    {
        "name": "4️⃣ Marketing",
        "full_name": "Marketing & Sales",
        "scenarios": [
            ("Customer Acquisition", "How do you get new customers?",
             [("Word of mouth only", 1), ("Multi-channel: social media, referrals, partnerships, ads", 3), ("I wait for them to find me", 0), ("I post on Facebook sometimes", 1)]),
            
            ("Value Proposition", "Why should customers choose you over competitors?",
             [("We're cheaper", 1), ("I can clearly explain our unique value", 3), ("I don't know", 0), ("We're better quality", 1)]),
            
            ("Customer Retention", "How do you keep existing customers coming back?",
             [("I don't - I focus on new customers", 0), ("Loyalty programs, great service, regular communication", 3), ("Low prices", 1), ("Hope they remember us", 0)]),
            
            ("Social Media", "How do you use social media for your business?",
             [("I don't", 0), ("Strategic posts, engage with customers, track what works", 3), ("Random posts when I remember", 1), ("I have a page but don't post", 0)]),
            
            ("Customer Feedback", "How do you collect and use customer feedback?",
             [("I don't ask for feedback", 0), ("I actively ask, track, and improve based on feedback", 3), ("Customers complain if they're unhappy", 1), ("I did a survey once", 1)]),
            
            ("Branding", "Do you have a consistent brand (logo, colors, messaging)?",
             [("No, I just use my business name", 0), ("Yes, and it's consistent across all touchpoints", 3), ("I have a logo my cousin made", 1), ("Branding is for big companies", 0)]),
            
            ("Sales Process", "What's your sales process?",
             [("Customer asks, I quote, they buy or don't", 1), ("Structured: lead capture, follow-up, proposal, close, after-sales", 3), ("I don't have a process", 0), ("I'm not good at sales", 0)]),
            
            ("Digital Presence", "Do you have a website or online presence?",
             [("No", 0), ("Yes - website, Google My Business, active social media", 3), ("Just a Facebook page", 1), ("I'm on WhatsApp", 1)]),
            
            ("Marketing Budget", "How much do you spend on marketing?",
             [("Nothing - I can't afford it", 0), ("5-10% of revenue, tracked for ROI", 3), ("Whatever I have left over", 1), ("I don't track it", 0)]),
            
            ("Customer Segmentation", "Do you know who your ideal customer is?",
             [("Anyone who will buy", 0), ("Yes - specific demographics, needs, behaviors", 3), ("Sort of - I know the type", 1), ("I've never thought about it", 0)])
        ]
    },
    {
        "name": "5️⃣ Operations",
        "full_name": "Operations & Technology",
        "scenarios": [
            ("Process Documentation", "Are your key business processes documented?",
             [("No, it's all in my head", 0), ("Yes, written down so anyone can follow", 3), ("Some are, most aren't", 1), ("I don't need to - I do everything", 0)]),
            
            ("Quality Control", "How do you ensure consistent quality?",
             [("I check everything myself", 1), ("Standard processes, checklists, spot checks", 3), ("Hope for the best", 0), ("Customers complain if there's a problem", 0)]),
            
            ("Inventory Management", "How do you manage stock/inventory?",
             [("I order when I run out", 1), ("System tracks stock levels, reorder points, turnover", 3), ("I don't sell physical products", 0), ("I keep lots of stock just in case", 0)]),
            
            ("Technology Adoption", "How do you use technology in your business?",
             [("Pen and paper mostly", 0), ("Accounting software, CRM, digital payments, cloud storage", 3), ("WhatsApp and Excel", 1), ("I'm not tech-savvy", 0)]),
            
            ("Supplier Relationships", "How do you manage supplier relationships?",
             [("I just order when I need stuff", 1), ("Regular communication, negotiate terms, have backups", 3), ("I use whoever is cheapest", 0), ("I have one supplier for everything", 1)]),
            
            ("Efficiency Improvement", "How often do you look for ways to work more efficiently?",
             [("Never - if it works, don't fix it", 0), ("Regularly review and optimize processes", 3), ("When something breaks", 1), ("I'm too busy to think about efficiency", 0)]),
            
            ("Data Security", "How do you protect business data?",
             [("I don't think about it", 0), ("Backups, passwords, cloud storage, POPIA compliance", 3), ("I save everything on my laptop", 1), ("What data?", 0)]),
            
            ("Automation", "Do you automate repetitive tasks?",
             [("No, I do everything manually", 0), ("Yes - invoicing, reminders, reports, social posts", 3), ("I've thought about it but haven't done it", 1), ("Automation is too expensive", 0)]),
            
            ("Capacity Planning", "How do you plan for busy periods?",
             [("I just work harder", 0), ("Forecast demand, plan resources, hire temp staff if needed", 3), ("Turn away customers", 0), ("Hope I can cope", 1)]),
            
            ("Continuous Improvement", "How do you stay updated on industry trends and best practices?",
             [("I don't have time", 0), ("Read industry news, attend workshops, network with peers", 3), ("Occasionally", 1), ("I rely on experience", 1)])
        ]
    }
]

# Create assessment sheets
for cap in capabilities:
    ws = wb.create_sheet(cap["name"])
    
    # Header
    ws['A1'] = f"{cap['full_name']} Assessment"
    ws['A1'].font = Font(size=16, bold=True, color=NAVY)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = center_align
    
    ws['A2'] = "Read each scenario and choose the response that best matches what YOU would do"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:E2')
    ws['A2'].alignment = center_align
    
    # Column headers
    headers = ['#', 'Scenario', 'Your Choice', 'Points', 'Feedback']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 40
    
    # Add scenarios
    row = 5
    for idx, (title, scenario, options) in enumerate(cap["scenarios"], start=1):
        # Question number and scenario
        ws.cell(row=row, column=1, value=idx).alignment = center_align
        ws.cell(row=row, column=2, value=f"{title}\n\n{scenario}").alignment = left_align
        
        # Dropdown for choices
        option_texts = [opt[0] for opt in options]
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(type="list", formula1=f'"{",".join(option_texts)}"', allow_blank=False)
        dv.add(ws.cell(row=row, column=3))
        ws.add_data_validation(dv)
        ws.cell(row=row, column=3).alignment = left_align
        
        # Points formula (VLOOKUP equivalent using IF statements)
        points_formula = "=IF(C{0}=\"\",0,".format(row)
        for opt_text, opt_points in options:
            points_formula += f'IF(C{row}="{opt_text}",{opt_points},'
        points_formula += "0" + ")" * len(options) + ")"
        
        ws.cell(row=row, column=4, value=points_formula).alignment = center_align
        
        # Feedback based on points
        feedback_formula = f'=IF(D{row}=3,"✅ Excellent!",IF(D{row}=2,"👍 Good",IF(D{row}=1,"⚠️ Could improve","❌ Needs attention")))'
        ws.cell(row=row, column=5, value=feedback_formula).alignment = center_align
        
        # Row height
        ws.row_dimensions[row].height = 80
        
        row += 1
    
    # Total score
    ws.cell(row=row+1, column=2, value="TOTAL SCORE:").font = Font(bold=True, size=12)
    ws.cell(row=row+1, column=2).alignment = Alignment(horizontal="right")
    total_formula = f"=SUM(D5:D{row-1})"
    ws.cell(row=row+1, column=4, value=total_formula).font = Font(bold=True, size=14, color=GREEN)
    ws.cell(row=row+1, column=4).alignment = center_align
    
    # Rating
    rating_formula = f'=IF(D{row+1}>=9,"🌟 STRONG",IF(D{row+1}>=7,"✅ COMPETENT",IF(D{row+1}>=4,"⚠️ DEVELOPING","🚨 URGENT")))'
    ws.cell(row=row+1, column=5, value=rating_formula).font = Font(bold=True, size=12)
    ws.cell(row=row+1, column=5).alignment = center_align

# ============================================================================
# SHEET 7: DASHBOARD
# ============================================================================
ws_dash = wb.create_sheet("📊 Dashboard", 0)  # Insert at beginning

ws_dash['A1'] = "Your Capability Compass Dashboard"
ws_dash['A1'].font = Font(size=18, bold=True, color=NAVY)
ws_dash.merge_cells('A1:F1')
ws_dash['A1'].alignment = center_align

ws_dash['A3'] = "Capability Area"
ws_dash['B3'] = "Score (out of 30)"
ws_dash['C3'] = "Rating"
ws_dash['D3'] = "Priority Actions"

for col in ['A', 'B', 'C', 'D']:
    ws_dash[f'{col}3'].fill = subheader_fill
    ws_dash[f'{col}3'].font = subheader_font
    ws_dash[f'{col}3'].alignment = center_align

# Link to other sheets
capability_names = ["Financial Management", "Strategic Planning", "Leadership & People", "Marketing & Sales", "Operations & Technology"]
sheet_refs = ["'1️⃣ Financial'", "'2️⃣ Strategic'", "'3️⃣ Leadership'", "'4️⃣ Marketing'", "'5️⃣ Operations'"]

for idx, (cap_name, sheet_ref) in enumerate(zip(capability_names, sheet_refs), start=4):
    ws_dash[f'A{idx}'] = cap_name
    ws_dash[f'A{idx}'].font = Font(bold=True)
    
    # Score formula
    ws_dash[f'B{idx}'] = f"={sheet_ref}!D16"  # Total from each sheet
    ws_dash[f'B{idx}'].alignment = center_align
    
    # Rating
    rating_formula = f'=IF(B{idx}>=27,"🌟 STRONG",IF(B{idx}>=21,"✅ COMPETENT",IF(B{idx}>=12,"⚠️ DEVELOPING","🚨 URGENT")))'
    ws_dash[f'C{idx}'] = rating_formula
    ws_dash[f'C{idx}'].alignment = center_align
    
    # Priority actions
    actions = {
        "Financial Management": "Review Operiva Cash Flow Forecast tool, attend bookkeeping workshop",
        "Strategic Planning": "Complete SWOT Analysis, set 90-day goals",
        "Leadership & People": "Document 3 key processes, delegate 1 task this week",
        "Marketing & Sales": "Define your ideal customer, create content calendar",
        "Operations & Technology": "Set up cloud backup, try 1 automation tool"
    }
    ws_dash[f'D{idx}'] = actions.get(cap_name, "")
    ws_dash[f'D{idx}'].alignment = left_align

# Set column widths
ws_dash.column_dimensions['A'].width = 25
ws_dash.column_dimensions['B'].width = 18
ws_dash.column_dimensions['C'].width = 18
ws_dash.column_dimensions['D'].width = 50

# Overall score
ws_dash['A10'] = "OVERALL CAPABILITY SCORE:"
ws_dash['A10'].font = Font(size=14, bold=True)
ws_dash['B10'] = "=SUM(B4:B8)"
ws_dash['B10'].font = Font(size=16, bold=True, color=GREEN)
ws_dash['B10'].alignment = center_align

ws_dash['A11'] = "Out of 150 points"
ws_dash['A11'].font = Font(size=10, italic=True)

# Interpretation
ws_dash['A13'] = "What Your Score Means:"
ws_dash['A13'].font = Font(size=12, bold=True)

interpretation = """
120-150: Strong owner-manager capabilities. Focus on continuous improvement.
90-119: Competent across most areas. Prioritize 1-2 development areas.
60-89: Developing. Significant gaps exist. Seek targeted support.
0-59: Urgent attention needed. Consider business coaching or mentorship.
"""

ws_dash['A14'] = interpretation
ws_dash['A14'].alignment = left_align
ws_dash.merge_cells('A14:D18')

# Next steps
ws_dash['A20'] = "🎯 Your Next Steps:"
ws_dash['A20'].font = Font(size=12, bold=True, color=NAVY)

next_steps = """
1. Review your scores in each capability area
2. Focus on areas marked "URGENT" or "DEVELOPING"
3. Download relevant Operiva artifacts for your weak areas
4. Set 3 specific goals for the next 90 days
5. Retake this assessment in 6 months to track progress
"""

ws_dash['A21'] = next_steps
ws_dash['A21'].alignment = left_align
ws_dash.merge_cells('A21:D26')

# Save workbook
output_path = '/home/ubuntu/Operiva/artifacts/03-Self-Assessment-Worksheet.xlsx'
wb.save(output_path)

print(f"✅ Self-Assessment Worksheet created: {output_path}")
print(f"   - 7 sheets: Instructions, 5 Capability Areas, Dashboard")
print(f"   - 50 scenario-based questions")
print(f"   - Automatic scoring and feedback")

