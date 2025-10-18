#!/usr/bin/env python3.11
"""
Create Cash Flow Forecasting Tool for SA SMMEs
Simplified 3-month rolling forecast template
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Define Operiva colors
NAVY = "1C2A4A"
GREEN = "27AE60"
LIGHT_BLUE = "E8F4F8"
LIGHT_GREEN = "E8F5E9"
YELLOW = "FFF9C4"
RED = "FFEBEE"

# Define styles
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
subheader_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
subheader_font = Font(color="FFFFFF", bold=True, size=11)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
currency_format = '#,##0.00'

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# SHEET 1: INSTRUCTIONS
# ============================================================================
ws_instructions = wb.create_sheet("📖 Instructions")

ws_instructions['A1'] = "Cash Flow Forecasting Tool for SA SMMEs"
ws_instructions['A1'].font = Font(size=18, bold=True, color=NAVY)
ws_instructions.merge_cells('A1:F1')

instructions_text = """
WHY CASH FLOW FORECASTING MATTERS:

Cash flow is the #1 reason SMMEs fail. You can be profitable on paper but run out of cash to pay suppliers, staff, or rent. This tool helps you:

✅ Predict cash shortages BEFORE they happen
✅ Plan for seasonal ups and downs
✅ Make better decisions about spending and hiring
✅ Sleep better at night!

HOW TO USE THIS TOOL:

STEP 1: Start with "Setup" tab
   • Enter your business name
   • Set your starting cash balance
   • Choose your forecast start month

STEP 2: Fill in "3-Month Forecast" tab
   • Enter expected cash IN (sales, loans, etc.)
   • Enter expected cash OUT (suppliers, salaries, rent, etc.)
   • The tool calculates your closing balance each month

STEP 3: Review "Dashboard" tab
   • See your cash flow trend (graph)
   • Identify months where cash is tight
   • Plan actions to avoid shortfalls

STEP 4: Update monthly
   • At month-end, compare actual vs forecast
   • Roll forward: delete oldest month, add new month
   • Adjust assumptions based on what you learned

SA-SPECIFIC TIPS:

🔌 Load Shedding: Budget for backup power (generator fuel, solar)
💰 Late Payments: SA average is 60-90 days - be conservative!
📈 Inflation: Costs rise 5-7% annually - factor this in
💳 VAT: If VAT-registered, remember you collect VAT but pay it to SARS later
🏦 Bank Charges: Don't forget transaction fees (can be R2000-5000/month)

COMMON MISTAKES TO AVOID:

❌ Being too optimistic about sales
❌ Forgetting irregular expenses (insurance, licenses, tax)
❌ Not planning for growth costs (more stock, more staff)
❌ Mixing business and personal cash

TIME REQUIRED: 30-45 minutes first time, 15 minutes monthly updates

LET'S GET STARTED! Go to the "Setup" tab →
"""

ws_instructions['A3'] = instructions_text
ws_instructions['A3'].alignment = left_align
ws_instructions.merge_cells('A3:F35')
ws_instructions.row_dimensions[3].height = 550

for col in range(1, 7):
    ws_instructions.column_dimensions[get_column_letter(col)].width = 15

# ============================================================================
# SHEET 2: SETUP
# ============================================================================
ws_setup = wb.create_sheet("⚙️ Setup")

ws_setup['A1'] = "Cash Flow Forecast Setup"
ws_setup['A1'].font = Font(size=16, bold=True, color=NAVY)
ws_setup.merge_cells('A1:D1')

# Business details
ws_setup['A3'] = "Business Name:"
ws_setup['A3'].font = Font(bold=True)
ws_setup['B3'] = "Your Business Name Here"
ws_setup['B3'].fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")

ws_setup['A4'] = "Forecast Start Month:"
ws_setup['A4'].font = Font(bold=True)
ws_setup['B4'] = date.today().replace(day=1)
ws_setup['B4'].number_format = 'MMMM YYYY'
ws_setup['B4'].fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")

ws_setup['A5'] = "Starting Cash Balance:"
ws_setup['A5'].font = Font(bold=True)
ws_setup['B5'] = 50000
ws_setup['B5'].number_format = currency_format
ws_setup['B5'].fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")

ws_setup['A7'] = "💡 TIP: Your starting cash balance is what's in your business bank account right now."
ws_setup['A7'].font = Font(italic=True, size=10)
ws_setup.merge_cells('A7:D7')

# Assumptions section
ws_setup['A10'] = "KEY ASSUMPTIONS (Update these based on your business)"
ws_setup['A10'].font = Font(size=14, bold=True, color=NAVY)
ws_setup.merge_cells('A10:D10')

assumptions = [
    ("Average Monthly Sales", 100000, "Conservative estimate - what you're confident you'll sell"),
    ("Sales Growth Rate", 0.05, "5% = 0.05 (monthly growth)"),
    ("% Cash Sales", 0.30, "30% = 0.30 (rest is credit)"),
    ("Average Collection Days", 60, "How long customers take to pay (SA average: 60-90 days)"),
    ("Cost of Goods Sold %", 0.50, "50% = 0.50 (if you sell R100, R50 is cost)"),
    ("Supplier Payment Days", 30, "How long before you pay suppliers"),
]

row = 12
ws_setup['A11'] = "Assumption"
ws_setup['B11'] = "Value"
ws_setup['C11'] = "Explanation"
for col in ['A', 'B', 'C']:
    ws_setup[f'{col}11'].fill = subheader_fill
    ws_setup[f'{col}11'].font = subheader_font

for assumption, value, explanation in assumptions:
    ws_setup[f'A{row}'] = assumption
    ws_setup[f'B{row}'] = value
    ws_setup[f'B{row}'].fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    if isinstance(value, float) and value < 1:
        ws_setup[f'B{row}'].number_format = '0.00%'
    elif isinstance(value, int) and value > 100:
        ws_setup[f'B{row}'].number_format = currency_format
    ws_setup[f'C{row}'] = explanation
    ws_setup[f'C{row}'].font = Font(italic=True, size=9)
    row += 1

ws_setup.column_dimensions['A'].width = 25
ws_setup.column_dimensions['B'].width = 15
ws_setup.column_dimensions['C'].width = 50

# ============================================================================
# SHEET 3: 3-MONTH FORECAST
# ============================================================================
ws_forecast = wb.create_sheet("📊 3-Month Forecast")

ws_forecast['A1'] = "3-Month Cash Flow Forecast"
ws_forecast['A1'].font = Font(size=16, bold=True, color=NAVY)
ws_forecast.merge_cells('A1:E1')

ws_forecast['A2'] = f"Business: [From Setup]    |    Forecast Period: [From Setup]"
ws_forecast['A2'].font = Font(italic=True)
ws_forecast.merge_cells('A2:E2')

# Column headers
ws_forecast['A4'] = "CASH FLOW ITEM"
ws_forecast['B4'] = "Month 1"
ws_forecast['C4'] = "Month 2"
ws_forecast['D4'] = "Month 3"
ws_forecast['E4'] = "3-Month Total"

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_forecast[f'{col}4'].fill = subheader_fill
    ws_forecast[f'{col}4'].font = subheader_font
    ws_forecast[f'{col}4'].alignment = center_align

# CASH IN section
row = 6
ws_forecast[f'A{row}'] = "💰 CASH IN (Money Coming In)"
ws_forecast[f'A{row}'].font = Font(bold=True, size=12, color=GREEN)
ws_forecast.merge_cells(f'A{row}:E{row}')

cash_in_items = [
    "Cash Sales",
    "Collections from Debtors (credit sales)",
    "Loans/Funding Received",
    "Other Income (specify)",
]

row += 1
for item in cash_in_items:
    ws_forecast[f'A{row}'] = item
    for col in ['B', 'C', 'D']:
        ws_forecast[f'{col}{row}'] = 0
        ws_forecast[f'{col}{row}'].number_format = currency_format
        ws_forecast[f'{col}{row}'].fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    ws_forecast[f'E{row}'] = f"=SUM(B{row}:D{row})"
    ws_forecast[f'E{row}'].number_format = currency_format
    row += 1

# Total Cash In
ws_forecast[f'A{row}'] = "TOTAL CASH IN"
ws_forecast[f'A{row}'].font = Font(bold=True)
for col in ['B', 'C', 'D', 'E']:
    start_row = row - len(cash_in_items)
    ws_forecast[f'{col}{row}'] = f"=SUM({col}{start_row}:{col}{row-1})"
    ws_forecast[f'{col}{row}'].number_format = currency_format
    ws_forecast[f'{col}{row}'].font = Font(bold=True)
    ws_forecast[f'{col}{row}'].fill = header_fill
    ws_forecast[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")

# CASH OUT section
row += 2
ws_forecast[f'A{row}'] = "💸 CASH OUT (Money Going Out)"
ws_forecast[f'A{row}'].font = Font(bold=True, size=12, color="FF0000")
ws_forecast.merge_cells(f'A{row}:E{row}')

cash_out_items = [
    "Suppliers/Stock Purchases",
    "Salaries & Wages",
    "Rent",
    "Utilities (electricity, water, internet)",
    "Transport/Fuel",
    "Marketing & Advertising",
    "Bank Charges & Interest",
    "Insurance",
    "Licenses & Permits",
    "Loan Repayments",
    "VAT Payment to SARS",
    "Tax (PAYE, Provisional Tax)",
    "Load Shedding Costs (generator, solar)",
    "Other Expenses (specify)",
]

row += 1
for item in cash_out_items:
    ws_forecast[f'A{row}'] = item
    for col in ['B', 'C', 'D']:
        ws_forecast[f'{col}{row}'] = 0
        ws_forecast[f'{col}{row}'].number_format = currency_format
        ws_forecast[f'{col}{row}'].fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    ws_forecast[f'E{row}'] = f"=SUM(B{row}:D{row})"
    ws_forecast[f'E{row}'].number_format = currency_format
    row += 1

# Total Cash Out
ws_forecast[f'A{row}'] = "TOTAL CASH OUT"
ws_forecast[f'A{row}'].font = Font(bold=True)
for col in ['B', 'C', 'D', 'E']:
    start_row = row - len(cash_out_items)
    ws_forecast[f'{col}{row}'] = f"=SUM({col}{start_row}:{col}{row-1})"
    ws_forecast[f'{col}{row}'].number_format = currency_format
    ws_forecast[f'{col}{row}'].font = Font(bold=True)
    ws_forecast[f'{col}{row}'].fill = header_fill
    ws_forecast[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")

# NET CASH FLOW
row += 2
ws_forecast[f'A{row}'] = "NET CASH FLOW (In - Out)"
ws_forecast[f'A{row}'].font = Font(bold=True, size=12)
total_in_row = 6 + len(cash_in_items) + 1
total_out_row = total_in_row + 2 + len(cash_out_items) + 1
for col in ['B', 'C', 'D', 'E']:
    ws_forecast[f'{col}{row}'] = f"={col}{total_in_row}-{col}{total_out_row}"
    ws_forecast[f'{col}{row}'].number_format = currency_format
    ws_forecast[f'{col}{row}'].font = Font(bold=True, size=12)

# OPENING BALANCE
row += 2
ws_forecast[f'A{row}'] = "Opening Cash Balance"
ws_forecast[f'B{row}'] = "=Setup!B5"  # From setup sheet
ws_forecast[f'B{row}'].number_format = currency_format
ws_forecast[f'C{row}'] = f"=D{row+1}"  # Previous month's closing
ws_forecast[f'C{row}'].number_format = currency_format
ws_forecast[f'D{row}'] = f"=E{row+1}"  # Previous month's closing
ws_forecast[f'D{row}'].number_format = currency_format

# CLOSING BALANCE
row += 1
ws_forecast[f'A{row}'] = "CLOSING CASH BALANCE"
ws_forecast[f'A{row}'].font = Font(bold=True, size=12, color=NAVY)
net_cash_row = row - 2
opening_row = row - 1
for col in ['B', 'C', 'D']:
    ws_forecast[f'{col}{row}'] = f"={col}{opening_row}+{col}{net_cash_row}"
    ws_forecast[f'{col}{row}'].number_format = currency_format
    ws_forecast[f'{col}{row}'].font = Font(bold=True, size=12, color=NAVY)
    ws_forecast[f'{col}{row}'].fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

# Warning if negative
row += 1
ws_forecast[f'A{row}'] = "⚠️ Cash Status"
for col_idx, col in enumerate(['B', 'C', 'D'], start=1):
    closing_cell = f"{col}{row-1}"
    ws_forecast[f'{col}{row}'] = f'=IF({closing_cell}<0,"🚨 NEGATIVE!",IF({closing_cell}<10000,"⚠️ Low","✅ Healthy"))'
    ws_forecast[f'{col}{row}'].alignment = center_align

ws_forecast.column_dimensions['A'].width = 35
for col in ['B', 'C', 'D', 'E']:
    ws_forecast.column_dimensions[col].width = 15

# ============================================================================
# SHEET 4: DASHBOARD
# ============================================================================
ws_dash = wb.create_sheet("📈 Dashboard", 0)  # Insert at beginning

ws_dash['A1'] = "Cash Flow Dashboard"
ws_dash['A1'].font = Font(size=18, bold=True, color=NAVY)
ws_dash.merge_cells('A1:F1')

ws_dash['A3'] = "Your 3-Month Cash Flow at a Glance"
ws_dash['A3'].font = Font(size=12, italic=True)
ws_dash.merge_cells('A3:F3')

# Summary table
ws_dash['A5'] = "Month"
ws_dash['B5'] = "Cash In"
ws_dash['C5'] = "Cash Out"
ws_dash['D5'] = "Net Flow"
ws_dash['E5'] = "Closing Balance"
ws_dash['F5'] = "Status"

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_dash[f'{col}5'].fill = subheader_fill
    ws_dash[f'{col}5'].font = subheader_font
    ws_dash[f'{col}5'].alignment = center_align

for month_idx in range(1, 4):
    row = 5 + month_idx
    col_letter = chr(66 + month_idx - 1)  # B, C, D
    
    ws_dash[f'A{row}'] = f"Month {month_idx}"
    ws_dash[f'B{row}'] = f"='📊 3-Month Forecast'!{col_letter}{total_in_row}"
    ws_dash[f'B{row}'].number_format = currency_format
    ws_dash[f'C{row}'] = f"='📊 3-Month Forecast'!{col_letter}{total_out_row}"
    ws_dash[f'C{row}'].number_format = currency_format
    ws_dash[f'D{row}'] = f"='📊 3-Month Forecast'!{col_letter}{net_cash_row}"
    ws_dash[f'D{row}'].number_format = currency_format
    ws_dash[f'E{row}'] = f"='📊 3-Month Forecast'!{col_letter}{opening_row+1}"
    ws_dash[f'E{row}'].number_format = currency_format
    ws_dash[f'F{row}'] = f"='📊 3-Month Forecast'!{col_letter}{opening_row+2}"
    ws_dash[f'F{row}'].alignment = center_align

# Key insights
ws_dash['A11'] = "🔍 KEY INSIGHTS"
ws_dash['A11'].font = Font(size=14, bold=True, color=NAVY)

insights = [
    ("Lowest Cash Month:", f"=IF(E6<E7,IF(E6<E8,\"Month 1\",\"Month 3\"),IF(E7<E8,\"Month 2\",\"Month 3\"))"),
    ("Biggest Expense:", "Check your Cash Out items in the Forecast tab"),
    ("Cash Runway:", "How many months can you survive with current cash?"),
]

row = 12
for label, value in insights:
    ws_dash[f'A{row}'] = label
    ws_dash[f'A{row}'].font = Font(bold=True)
    ws_dash[f'B{row}'] = value
    ws_dash.merge_cells(f'B{row}:D{row}')
    row += 1

# Action items
ws_dash['A16'] = "🎯 ACTION ITEMS"
ws_dash['A16'].font = Font(size=14, bold=True, color=NAVY)

actions = """
✅ If all months are positive: Great! Consider investing surplus or building emergency fund.

⚠️ If any month is low (<R10,000): 
   • Chase late-paying customers
   • Negotiate payment terms with suppliers
   • Delay non-essential spending

🚨 If any month is negative:
   • URGENT: Arrange overdraft or short-term loan
   • Cut costs immediately
   • Push for faster customer payments
   • Consider delaying supplier payments (communicate with them!)

💡 Monthly Review Checklist:
   □ Compare actual vs forecast - where were you wrong?
   □ Update next month's forecast based on what you learned
   □ Check if you need to adjust prices or cut costs
   □ Ensure you have at least 1 month's expenses in reserve
"""

ws_dash['A17'] = actions
ws_dash['A17'].alignment = left_align
ws_dash.merge_cells('A17:F30')
ws_dash.row_dimensions[17].height = 300

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_dash.column_dimensions[col].width = 15

# Save workbook
output_path = '/home/ubuntu/Operiva/artifacts/03-Cash-Flow-Forecast.xlsx'
wb.save(output_path)

print(f"✅ Cash Flow Forecast created: {output_path}")
print(f"   - 4 sheets: Dashboard, Instructions, Setup, 3-Month Forecast")
print(f"   - SA-specific guidance (load shedding, VAT, late payments)")
print(f"   - Simple and actionable for SMMEs")

